#!/usr/bin/env python3
"""
refresh-data.py
Converts services.xlsx (all sheets) and internet_offers.xlsx → api/data/*.json
Run this whenever you update either spreadsheet:
  python3 scripts/refresh-data.py
"""

import sys
import json
import pathlib

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed. Run: pip3 install openpyxl")

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT  = ROOT / "api" / "data"
OUT.mkdir(parents=True, exist_ok=True)


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v)


# Each sheet has different column names; map them to a common schema:
#   site, address, phone, email, offerings, source, sheet
SHEET_MAPS = {
    "master list": {
        "site":      ["Organization"],
        "address":   ["Address"],
        "phone":     ["Phone"],
        "email":     ["Email"],
        "offerings": ["Reource Offering", "Resource Description"],  # note typo in sheet
        "source":    ["Source"],
    },
    "digital skill training": {
        "site":      ["Site"],
        "address":   ["Address"],
        "phone":     ["Phone"],
        "email":     ["Email"],
        "offerings": ["Programming and Resource Offerings"],
        "source":    ["Website"],
    },
    "device access resources": {
        "site":      ["Site"],
        "address":   ["Address"],
        "phone":     ["Phone"],
        "email":     ["Email"],
        "offerings": ["Programming and Resource Offerings"],
        "source":    ["Source"],
    },
    "skills & device access": {
        "site":      ["Organization"],
        "address":   ["Address"],
        "phone":     ["Phone"],
        "email":     ["Email"],
        "offerings": ["Programming and Resource Offerings"],
        "source":    ["Source"],
    },
    "national programs": {
        "site":      ["Organization"],
        "address":   [],
        "phone":     [],
        "email":     [],
        "offerings": ["Resource", "Content Focus"],
        "source":    ["Source"],
    },
    "workforce programs": {
        "site":      ["Site"],
        "address":   ["Address"],
        "phone":     ["Phone"],
        "email":     ["Email"],
        "offerings": ["Programming and Resource Offerings"],
        "source":    ["Source"],
    },
}


def first_value(row_dict: dict, candidates: list[str]) -> str:
    for key in candidates:
        v = row_dict.get(key, "")
        if v:
            return v
    return ""


def load_services(src: pathlib.Path) -> list[dict]:
    wb = openpyxl.load_workbook(src)
    all_rows = []

    for sheet_name in wb.sheetnames:
        mapping = SHEET_MAPS.get(sheet_name)
        if mapping is None:
            print(f"    (no mapping for sheet '{sheet_name}', skipping)")
            continue

        ws = wb[sheet_name]
        headers = [cell_str(cell.value) for cell in ws[1]]
        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw = {headers[i]: cell_str(row[i]) for i in range(len(headers))}

            site = first_value(raw, mapping["site"])
            if not site:
                continue  # skip blank rows

            # Combine multi-field offerings with a separator
            offering_parts = [first_value(raw, [k]) for k in mapping["offerings"]]
            offerings = " — ".join(p for p in offering_parts if p)

            all_rows.append({
                "Site":      site,
                "Address":   first_value(raw, mapping["address"]),
                "Phone":     first_value(raw, mapping["phone"]),
                "Email":     first_value(raw, mapping["email"]),
                "Programming and Resource Offerings": offerings,
                "Source":    first_value(raw, mapping["source"]),
                "Sheet":     sheet_name,
            })
            count += 1

        print(f"    {sheet_name}: {count} rows")

    return all_rows


def xlsx_to_json(src: pathlib.Path) -> list[dict]:
    """Generic single-sheet loader (used for internet_offers)."""
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    headers = [cell_str(cell.value) for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        obj = {headers[i]: cell_str(row[i]) for i in range(len(headers)) if headers[i]}
        rows.append(obj)
    return rows


# ── services.xlsx ─────────────────────────────────────────────────────────────
src = ROOT / "public" / "services.xlsx"
if src.exists():
    data = load_services(src)
    dst = OUT / "services.json"
    dst.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    print(f"  OK    services.xlsx → services.json  ({len(data)} total rows across all sheets)")
else:
    print(f"  SKIP  services.xlsx not found")

# ── internet_offers.xlsx ──────────────────────────────────────────────────────
src = ROOT / "public" / "internet_offers.xlsx"
if src.exists():
    data = xlsx_to_json(src)
    dst = OUT / "internet_offers.json"
    dst.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    print(f"  OK    internet_offers.xlsx → internet_offers.json  ({len(data)} rows)")
else:
    print(f"  SKIP  internet_offers.xlsx not found")

print("Done.")
